"""
Workbook writer for the AXIM property due-diligence tool.
Produces a clean, branded multi-tab .xlsx from a run_property() result dict.

Tabs:
  Summary          headline answer to every item on Jake's list, grouped, with source
  Assessments      DOF market value & assessed value, 5+ years
  Benefits         exemptions / abatements (business & construction)
  Violations       open DOB + ECB/environmental + HPD violations
  ACRIS            mortgage / deed / owner + full recorded-document list
  Tax & Account    Statement of Account link + parsed balance
  Raw_PLUTO        full PLUTO row (debugging / extra fields)
  Run_Log          per-source status, row count, timing, errors
"""

import re as _re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Strip characters that openpyxl/Excel rejects (control chars, etc.)
_ILLEGAL = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _safe(v):
    if isinstance(v, str):
        return _ILLEGAL.sub('', v)
    return v

NAVY = "1F3864"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
RED = "F4CCCC"
GREEN = "D9EAD3"

H1 = Font(bold=True, color="FFFFFF", size=13)
H2 = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def _section(ws, row, title, span=2):
    c = ws.cell(row=row, column=1, value=title)
    c.font = H2
    c.fill = _fill(NAVY)
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = _fill(NAVY)
    return row + 1


def _kv(ws, row, key, val):
    a = ws.cell(row=row, column=1, value=key)
    a.font = BOLD
    a.fill = _fill(GREY)
    a.alignment = TOP
    a.border = BORDER
    b = ws.cell(row=row, column=2, value=("" if val is None else val))
    b.alignment = WRAP
    b.border = BORDER
    return row + 1


def _table(ws, start_row, headers, rows, money_cols=()):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = H2
        c.fill = _fill(NAVY)
        c.alignment = WRAP
        c.border = BORDER
    r = start_row + 1
    for rec in rows:
        for j, h in enumerate(headers, start=1):
            v = rec.get(h)
            c = ws.cell(row=r, column=j, value=("" if v is None else _safe(v)))
            c.alignment = WRAP
            c.border = BORDER
            if h in money_cols and isinstance(v, (int, float)):
                c.number_format = '$#,##0'
        r += 1
    return r


def _autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_workbook(res, path):
    ident = res["ident"]
    wb = Workbook()

    # ---------- Summary ----------
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    title = ws.cell(row=1, column=1,
                    value="AXIM Real Estate Partners — Property Due-Diligence Summary")
    title.font = H1
    title.fill = _fill(NAVY)
    ws.cell(row=1, column=2).fill = _fill(NAVY)
    ws.merge_cells("A1:B1")

    r = 3
    r = _kv(ws, r, "Input", ident.get("input"))
    r = _kv(ws, r, "Address", ident.get("address") or res["pluto_sum"].get("PLUTO address"))
    boro = ident.get("boro")
    from axim_property_dd import BOROUGH_NAME
    r = _kv(ws, r, "Borough", f"{BOROUGH_NAME.get(boro, boro)} ({boro})")
    r = _kv(ws, r, "BBL", ident.get("bbl"))
    r = _kv(ws, r, "Block / Lot", f"{ident.get('block')} / {ident.get('lot')}")
    r = _kv(ws, r, "BIN", ident.get("bin"))
    r += 1

    r = _section(ws, r, "ZoLa — Zoning, Lot, Units, Age, Size")
    for k, v in res["pluto_sum"].items():
        r = _kv(ws, r, k, v)
    r += 1

    latest = res["assess"][0] if res["assess"] else {}
    r = _section(ws, r, "DOF — Valuation (most recent; full history on Assessments tab)")
    r = _kv(ws, r, "Tax year", latest.get("Tax year"))
    r = _kv(ws, r, "Market value (est.)",
            f"${latest.get('Market value (est.)'):,.0f}" if latest.get("Market value (est.)") else "")
    r = _kv(ws, r, "Assessed total",
            f"${latest.get('Assessed total'):,.0f}" if latest.get("Assessed total") else "")
    r = _kv(ws, r, "Benefits / exemptions",
            "; ".join(b.get("Benefit / holder", "") for b in res["benefits"]
                      if b.get("Benefit / holder")) or "None found")
    r += 1

    r = _section(ws, r, "DOB — Certificate of Occupancy & Violations")
    if res["cofo_sum"]:
        for k, v in res["cofo_sum"].items():
            r = _kv(ws, r, k, v)
    else:
        r = _kv(ws, r, "Certificate of Occupancy",
                "No C of O on file with DOB (common for pre-1938 buildings)")
    r = _kv(ws, r, "Open DOB violations", len(res["dobv"]))
    r = _kv(ws, r, "Open ECB/environmental violations", len(res["ecbv"]))
    r = _kv(ws, r, "Open HPD violations", len(res["hpdv"]))
    r += 1

    r = _section(ws, r, "ACRIS — Mortgage, Deed, Owner")
    if res["acris_sum"]:
        for k, v in res["acris_sum"].items():
            r = _kv(ws, r, k, v)
    else:
        r = _kv(ws, r, "Recorded transactions",
                "No private market transactions on file (likely city / government / long-held)")
    r += 1

    r = _section(ws, r, "DOF — Tax Bill & Account Balance")
    if res["tax_info"]:
        for k, v in res["tax_info"].items():
            r = _kv(ws, r, k, v)
        # Statement located but no amount due / market value parsed → likely tax-exempt
        amt = res["tax_info"].get("Amount due")
        mv  = res["tax_info"].get("Est. market value (from bill)")
        if (res["tax_info"].get("Statement found") in ("Yes", True)
                and not amt and not mv):
            r = _kv(ws, r, "Note",
                    "Statement located but no amount due / market value parsed — "
                    "typical for tax-exempt owners (DOE, NYCHA, non-profit, government)")
    else:
        r = _kv(ws, r, "Tax bill", "No DOF Statement of Account found on probed dates")

    _autosize(ws, {"A": 34, "B": 70})

    # ---------- Assessments ----------
    ws = wb.create_sheet("Assessments")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="DOF Market Value & Assessment History").font = BOLD
    headers = ["Tax year", "Tax class", "Market value (est.)", "Assessed land",
               "Assessed total", "Exempt total", "Bldg class"]
    _table(ws, 3, headers, res["assess"],
           money_cols={"Market value (est.)", "Assessed land", "Assessed total",
                       "Exempt total"})
    _autosize(ws, {get_column_letter(i): 18 for i in range(1, len(headers) + 1)})

    # ---------- Benefits ----------
    ws = wb.create_sheet("Benefits")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="DOF Benefits — Exemptions & Abatements").font = BOLD
    headers = ["Type", "Benefit / holder", "Code", "Tax year", "Amount"]
    rows = res["benefits"] or [{"Type": "", "Benefit / holder": "None found on DOF datasets"}]
    _table(ws, 3, headers, rows, money_cols={"Amount"})
    _autosize(ws, {"A": 14, "B": 40, "C": 12, "D": 12, "E": 16})

    # ---------- Violations ----------
    ws = wb.create_sheet("Violations")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Open Violations — DOB / ECB / HPD").font = BOLD
    all_v = res["dobv"] + res["ecbv"] + res["hpdv"]
    headers = ["Agency", "Class", "Type", "Issued", "Status", "Description",
               "Number", "Penalty due"]
    rows = all_v or [{"Agency": "", "Description": "No open violations found"}]
    _table(ws, 3, headers, rows)
    _autosize(ws, {"A": 10, "B": 8, "C": 22, "D": 12, "E": 16, "F": 55,
                   "G": 16, "H": 14})

    # ---------- ACRIS ----------
    ws = wb.create_sheet("ACRIS")
    ws.sheet_view.showGridLines = False
    r = 1
    ws.cell(row=r, column=1, value="ACRIS — Current Mortgage / Deed / Owner").font = BOLD
    r += 2
    if res["acris_sum"]:
        for k, v in res["acris_sum"].items():
            _kv(ws, r, k, v)
            r += 1
    else:
        _kv(ws, r, "Status",
            "No private market transactions on file (likely city / government / long-held)")
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="All recorded documents (most recent first)").font = BOLD
    r += 1
    if res["acris_txns"]:
        _table(ws, r, ["Doc type", "Doc date", "Recorded", "Amount", "Doc ID"],
               res["acris_txns"], money_cols={"Amount"})
    else:
        _kv(ws, r, "Status", "No recorded documents found")
    _autosize(ws, {"A": 34, "B": 16, "C": 14, "D": 16, "E": 30})

    # ---------- Raw PLUTO ----------
    ws = wb.create_sheet("Raw_PLUTO")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Field").font = BOLD
    ws.cell(row=1, column=2, value="Value").font = BOLD
    rr = 2
    for k, v in (res.get("pluto_raw") or {}).items():
        ws.cell(row=rr, column=1, value=k)
        ws.cell(row=rr, column=2, value=str(v))
        rr += 1
    _autosize(ws, {"A": 26, "B": 50})

    # ---------- Run Log ----------
    ws = wb.create_sheet("Run_Log")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Source retrieval log").font = BOLD
    hdr = ["Source", "Status", "Rows", "Seconds", "Error"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = H2
        c.fill = _fill(NAVY)
    rr = 4
    for name, status, n, secs, err in res["log"]:
        ws.cell(row=rr, column=1, value=name)
        sc = ws.cell(row=rr, column=2, value=status)
        sc.fill = _fill(GREEN if status == "OK" else RED)
        ws.cell(row=rr, column=3, value=n)
        ws.cell(row=rr, column=4, value=secs)
        ws.cell(row=rr, column=5, value=err)
        rr += 1
    import datetime
    ws.cell(row=rr + 1, column=1,
            value=f"Generated {datetime.datetime.now():%Y-%m-%d %H:%M}").font = \
        Font(italic=True, color="808080")
    _autosize(ws, {"A": 22, "B": 10, "C": 8, "D": 10, "E": 70})
    wb.save(path)
    return path
